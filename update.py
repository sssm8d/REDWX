import requests
import os
from openpyxl import load_workbook
import pandas as pd
import datetime

fmt_acct = u'$#,##0.00;[Red]$(#,##0.00);-;@'

def getYTDinfo():
    #works for the rest of 2020
    url = 'https://query1.finance.yahoo.com/v7/finance/download/REDWX?period1=1577750400&period2=1609372800&interval=1d&events=history'

    myfile = requests.get(url, allow_redirects=True)

    #print( "%s\REDWX.csv" % (os.getcwd()) )
    file = open("%s\REDWX.csv" % (os.getcwd()), 'wb')
    file.write(myfile.content)
    file.close()
    
def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    #Loops through selected Rows
    for i in range(startRow,endRow + 1,1):
        #Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol,endCol+1,1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
        #Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)
 
    return rangeSelected
    
def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving, copiedData):
    countRow = 0
    for i in range(startRow,endRow + 1,1):
        countCol = 0
        for j in range(startCol,endCol+1,1):
            if ( (j is startCol) and (i is not startRow) ):
                if ( copiedData[countRow][countCol] is not None ):
                    #print( str(copiedData[countRow][countCol]) )
                    copiedData[countRow][countCol] = datetime.datetime.strptime( str(copiedData[countRow][countCol]) , "%Y-%m-%d")
                sheetReceiving.cell(row = i, column = j).number_format = 'YYYY-MM-DD'
            else:
                sheetReceiving.cell(row = i, column = j).number_format = fmt_acct
     
            sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
            #if ( (j is startCol) and (i is not startRow) ):
            #copiedData[countRow][countCol] = datetime.datetime.strptime(str(copiedData[countRow][countCol]), "%Y-%m-%d")
                
            countCol += 1
        countRow += 1

def updateDataSheet():
    import_data = pd.read_csv( "REDWX.csv" )
    import_data.to_excel(r'REDWX_data.xlsx', sheet_name='Data', index = False)

    workbookin = load_workbook(filename = 'REDWX_data.xlsx')
    data_sheetin = workbookin["Data"]
    for i in range( 2, 400 ):
        if ( data_sheetin.cell(row = i, column = 1).value is not None ):
            data_sheetin.cell(row = i, column = 1).value = datetime.datetime.strptime( str(data_sheetin.cell(row = i, column = 1).value) , "%Y-%m-%d")
        data_sheetin.cell(row = i, column = 1).number_format = 'YYYY-MM-DD'
        
    workbookin.save('REDWX_data.xlsx')

    os.remove("REDWX.csv")

#get info from internet
getYTDinfo()

# save and close REDWX.xlsx
updateDataSheet()

# open dashboard
filename  = "%s\REDWX.xlsm" % (os.getcwd())
os.startfile(filename)

#make into button on dashboard